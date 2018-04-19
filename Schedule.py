#import xlrd
#import xlwt
import time
import openpyxl

localtime = time.asctime(time.localtime(time.time()) )
print (localtime)
day = localtime.split(' ', 1)[0]
#print (day)

'''
workbook = xlrd.open_workbook('D:/Smart Mirror/SmartMirror/Time Tables/Schedule.xlsx')
worksheet = workbook.sheet_by_name(day)
print (workbook.sheet_names())
print (worksheet.row_values(0))

cells = worksheet.row_slice(rowx=1, start_colx=0, end_colx=2)
for cell in cells:
    print (cell.value)
    '''
workbook = openpyxl.load_workbook('D:/Smart Mirror/SmartMirror/Time Tables/Schedule.xlsx')
worksheet = workbook[day]


currenttime = localtime.split(' ')[3]
hour = currenttime.split(':')[0]
minute = currenttime.split(':')[1]

hour = int(hour)
minute = int(minute)
print(currenttime)
print(hour)
print(minute)

if hour <= 10 and minute <= 15:
    for row in worksheet.iter_rows(min_row=2, min_col=3, max_row=50, max_col=3):
        for cell in row:
        #    print(cell.value, end=" ")
        #print()
            if cell.value == 'Yes':
                print('unavailable')
            elif cell.value == 'No':
                print('Available')
            elif cell.value== ' ':
                break    
elif hour == 11 and minute >= 15:
    for row in worksheet.iter_rows(min_row=2, min_col=4, max_row=50, max_col=4):
        for cell in row:
        #    print(cell.value, end=" ")
        #print()
            if cell.value == 'Yes':
                print('unavailable')
            elif cell.value == 'No':
                print('Available')
            elif cell.value== ' ':
                break 
elif hour <= 12 and minute <= 15:
    for row in worksheet.iter_rows(min_row=2, min_col=5, max_row=50, max_col=5):
        for cell in row:
        #    print(cell.value, end=" ")
        #print()
            if cell.value == 'Yes':
                print('unavailable')
            elif cell.value == 'No':
                print('Available')
            elif cell.value== ' ':
                break   